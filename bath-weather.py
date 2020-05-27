import pandas as pd 
#dataframes will be created later on

import requests 

#requests are required to get data from a given URL

from bs4 import BeautifulSoup

import numpy as np 

import matplotlib.pyplot as plt 
#we can eventually plot temperatures on a graph


page = requests.get('https://www.metoffice.gov.uk/weather/forecast/gcnk62de6#?date=2020-05-27')
#page we are interested in 

soup = BeautifulSoup(page.content, 'html.parser') #html parser parses through the page we have requested and stores it under soup

week = soup.find(id='dayNav') #weekly data is found under the 'dayNav id'

days = week.find_all(class_='tab-day') #we want the days of the week, found under the div class 'tab-day', returns a list 

high_temp = week.find_all(class_='tab-temp-high') #we want projected high temps for the given days, found under div class 'tab-temp-high'

low_temp = week.find_all(class_='tab-temp-low') #we want projected low temps for the given days, found under div class 'tab-temp-low'

summary = week.find_all(class_='summary-text hide-xs-only') #there is also a summmary text for each day

#The data found under each class is stored in each respective list, but we want the text only, so to remove the code:

day_titles = [day.get_text().replace('\n','') for day in days] #returns ['Day1','Day2','Day3'....'Day n']

high_temperatures = [temp.get_text() for temp in high_temp] #returns [i,j,k....n]

low_temperatures = [temp.get_text() for temp in low_temp] #returns [i,j,k....n]

summary_stat = [item.get_text().replace('\n','') for item in summary] #returns summary string for each corresponding day


weather_stuff = pd.DataFrame(
	{'Days':day_titles,'High':high_temperatures,
	'Low': low_temperatures, 'Summary': summary_stat})

#uses pandas to create data frame.  each key is a column heading, corresponding values go under heading

#create a blank excel file titled "Bath Weather.xlsx" to store results

from openpyxl import load_workbook #loads excel workbook 
from openpyxl.utils.dataframe import dataframe_to_rows #exports data frame to excel

file_location = r'/Users/amitkamlani/Desktop/Bath Weather.xlsx' #file path

write_location = r'/Users/amitkamlani/Desktop/Bath Weather.xlsx' #location we want to write to: same as file path as we want to append the file every time code is run


book = load_workbook(file_location) #blank excel workbook we created

sheet = book.create_sheet("Weekly Data") #creates new sheet

for row in dataframe_to_rows(weather_stuff,index=False, header=True):
	sheet.append(row) #every time code is run, dataframe is appended to a new excel sheet

book.save(write_location) #book is saved

